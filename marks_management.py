from tkinter import *
from tkinter import ttk
from PIL import Image,ImageTk
from tkinter import messagebox as msg
import mysql.connector as conn
from tkinter import filedialog
from openpyxl import Workbook

win1=Tk()
win1.title("WELCOME")
win1.resizable(True,True)
win1.configure(bg="#FDFEFE")

#cursor.execute("create table users(username varchar(25) primary key,email varchar(30) not null,password varchar(25) not null,schoolname varchar(50) not null,profilephoto varchar(150) not null)")

##Backgroung Image
#bg1=Image.open("C:\\Users\\DELL\\Desktop\\student_database\\bg1.jpg")
#bg=ImageTk.PhotoImage(bg1)
#bg_label=Label(win1,image=bg)
#bg_label.pack()
##Background Image Ends

main_win1_label=Label(win1,text="MARKS MANAGEMENT",font=("Algerian",60),bg="#FDFEFE",foreground="#FF5733")
main_win1_label.place(x=390,y=0)

#Username label and entry starts
username_win1_label=Label(win1,text="USERNAME",width=12,bg="#3E9DE4",foreground="#FDFEFE",font=("Algerian",40),relief="raised",borderwidth=10)
username_win1_label.place(x=250,y=180)

username_win1_entry=StringVar()
username_win1_entry=Entry(win1,width=15,textvariable=username_win1_entry,font=("bold",45),relief="raised",bd=5,bg="#FDFEFE")
username_win1_entry.place(x=750,y=180)
username_win1_entry.focus()
#Username closes


#Email label and entry starts
email_win1_label=Label(win1,text="Email",width=12,bg="#3E9DE4",foreground="#FDFEFE",font=("Algerian",40),relief="raised",borderwidth=10)
email_win1_label.place(x=250,y=280)

email_win1_entry=StringVar()
email_win1_entry=Entry(win1,width=15,textvariable=email_win1_entry,font=("bold",45),relief="raised",bd=5,bg="#FDFEFE")
email_win1_entry.place(x=750,y=280)
email_win1_entry.focus()
#Email closes


#Password(pass) label and entry starts
pass_win1_label=Label(win1,text="Password",width=12,bg="#3E9DE4",foreground="#FDFEFE",font=("Algerian",40),relief="raised",borderwidth=10)
pass_win1_label.place(x=250,y=380)

pass_win1_entry=StringVar()
pass_win1_entry=Entry(win1,width=15,textvariable=pass_win1_entry,font=("bold",45),relief="raised",bd=5,bg="#FDFEFE",show="*")
pass_win1_entry.place(x=750,y=380)
pass_win1_entry.focus()
#Password(pass) closes

#School Name label and entry starts
school_win1_label=Label(win1,text="SCHOOL NAME",width=12,bg="#3E9DE4",foreground="#FDFEFE",font=("Algerian",40),relief="raised",borderwidth=10)
school_win1_label.place(x=250,y=480)

school_win1_entry=StringVar()
school_win1_entry=Entry(win1,width=15,textvariable=school_win1_entry,font=("bold",45),relief="raised",bd=5,bg="#FDFEFE")
school_win1_entry.place(x=750,y=480)
school_win1_entry.focus()
#School Name closes





#Register/Login Button
def log1_click():
    win1.destroy()
    win2=Tk()
    win2.title("LOGIN")
    win2.configure(bg="#FDFEFE")
    main_win2_label=Label(win2,text="MARKS MANAGEMENT",font=("Algerian",60),bg="#FDFEFE",foreground="#FF5733")
    main_win2_label.place(x=390,y=0)
    #Username label and entry starts
    username_win2_label=Label(win2,text="USERNAME",width=12,bg="#3E9DE4",foreground="#FDFEFE",font=("Algerian",40),relief="raised",borderwidth=10)
    username_win2_label.place(x=250,y=180)

    username_win2_entry=StringVar()
    username_win2_entry=Entry(win2,width=15,textvariable=username_win2_entry,font=("bold",45),relief="raised",bd=5,bg="#FDFEFE")
    username_win2_entry.place(x=750,y=180)
    username_win2_entry.focus()
    #Username closes


    #Password(pass) label and entry starts
    pass_win2_label=Label(win2,text="Password",width=12,bg="#3E9DE4",foreground="#FDFEFE",font=("Algerian",40),relief="raised",borderwidth=10)
    pass_win2_label.place(x=250,y=380)

    pass_win2_entry=StringVar()
    pass_win2_entry=Entry(win2,width=15,textvariable=pass_win2_entry,font=("bold",45),relief="raised",bd=5,bg="#FDFEFE",show="*")
    pass_win2_entry.place(x=750,y=380)
    #Password(pass) closes
    

    def reg_back_click():
        #win1.open()
        pass
    win2_log_button=Button(win2,text="Not a Member? Register",bg="#FDFEFE",foreground="blue",command=reg_back_click,relief="raised")
    win2_log_button.place(x=800,y=480)

    #login_cloud
    def log2_click():
        cursor.execute("select * from users")
        cred_get=cursor.fetchall()
        log_username_get=username_win2_entry.get()
        log_pass_get=pass_win2_entry.get()
        
        for fetch in cred_get:
            if fetch[0]==log_username_get:
                log_email_get=fetch[1]
                if fetch[2]==log_pass_get:
                    auth=True
                else:
                    auth=False
                    msg.showwarning("Warning","Please Enter Correct Credentials")
        if auth==True:
            win2.destroy()
            dash=Tk()
            dash.title("WELCOME  " + log_username_get.upper())
            
            frame_label=Label(dash,text='',width=212,relief='groove',height=11,bd=20)
            frame_label.place(x=0,y=0)
            
            cursor.execute("select * from users")
            fetcher=cursor.fetchall()
            for photo_fetch in fetcher:
                if photo_fetch[0]==log_username_get:
                    photo_fetch_address=photo_fetch[-1]
                    school_fetch_name=photo_fetch[-2]
                    #print(photo_fetch_address)
            open_image=Image.open(photo_fetch_address)
            dash_photo=ImageTk.PhotoImage(open_image)
            photo_label=Label(dash,image=dash_photo)
            photo_label.place(x=20,y=20)

            dash_school_label=Label(dash,text=school_fetch_name,font=("Algerian",40))
            dash_school_label.place(x=150,y=60)

            manage_label=Label(dash,text='',width=140,height=35,bd=20,relief='groove')
            manage_label.place(x=500,y=210)

            manage_heading=Label(dash,text='Manage Data',width=28,height=2,font=("Algerian",40),relief='raised')
            manage_heading.place(x=540,y=230)
            def update_data():
                global roll_update
                global roll_entry_update
                global class_update
                global class_entry_update
                global english_update
                global english_entry_update
                global t_h_entry_update
                global t_h_entry1_update
                global science_update
                global science_entry_update
                global Social_update
                global Social_entry_update
                global computer_update
                global computer_entry_update
                global exam_label_update
                global exam_update_combo
                global maths_update
                global maths_entry_update
                
                roll_update=Label(dash,text='Roll.No',width=10,font=("bold",20))
                roll_update.place(x=540,y=460)
                roll_entry_update_var=StringVar()
                roll_entry_update=Entry(dash,textvariable=roll_entry_update_var,width=10,font=("bold",20))
                roll_entry_update.place(x=700,y=460)


                class_update=Label(dash,text='Class',width=10,font=("bold",20))
                class_update.place(x=540,y=500)
                class_entry_update_var=StringVar()
                class_entry_update=Entry(dash,textvariable=class_entry_update_var,width=10,font=("bold",20))
                class_entry_update.place(x=700,y=500)


                english_update=Label(dash,text='English',width=10,font=("bold",20))
                english_update.place(x=540,y=540)
                english_entry_update_var=StringVar()
                english_entry_update=Entry(dash,textvariable=english_entry_update_var,width=10,font=("bold",20))
                english_entry_update.place(x=700,y=540)


                t_h_entry_update_var=StringVar()
                t_h_entry_update=ttk.Combobox(dash,textvariable=t_h_entry_update_var,font=("bold",15),width=5)
                t_h_entry_update['values']=('telugu','hindi')
                t_h_entry_update.current(0)
                t_h_entry_update.place(x=540,y=580)

                t_h_entry1_update_var=StringVar()
                t_h_entry1_update=Entry(dash,textvariable=t_h_entry1_update_var,width=10,font=("bold",20))
                t_h_entry1_update.place(x=700,y=580)


                
                science_update=Label(dash,text='Science',width=10,font=("bold",20))
                science_update.place(x=538,y=610)
                science_entry_update_var=StringVar()
                science_entry_update=Entry(dash,textvariable=science_entry_update_var,width=10,font=("bold",20))
                science_entry_update.place(x=700,y=620)


                Social_update=Label(dash,text='Social',width=10,font=("bold",20))
                Social_update.place(x=534,y=660)
                Social_entry_update_var=StringVar()
                Social_entry_update=Entry(dash,textvariable=Social_entry_update_var,width=10,font=("bold",20))
                Social_entry_update.place(x=700,y=660)


                computer_update=Label(dash,text='Computer',width=10,font=("bold",20))
                computer_update.place(x=534,y=690)
                computer_entry_update_var=StringVar()
                computer_entry_update=Entry(dash,textvariable=computer_entry_update_var,width=10,font=("bold",20))
                computer_entry_update.place(x=700,y=700)


                

                exam_label_update=Label(dash,text='Exam',width=10,font=("bold",20))
                exam_label_update.place(x=850,y=500)#850,500
                exam_update_var=StringVar()
                exam_update_combo=ttk.Combobox(dash,textvariable=exam_update_var,width=20,font=("bold",20))
                exam_update_combo['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                exam_update_combo.current(0)#1000,510
                exam_update_combo.place(x=1000,y=510)


                maths_update=Label(dash,text='Maths',width=10,font=("bold",20))
                maths_update.place(x=850,y=460)#850,460
                maths_entry_update_var=StringVar()
                maths_entry_update=Entry(dash,textvariable=maths_entry_update_var,width=10,font=("bold",20))
                maths_entry_update.place(x=1000,y=460)#1000,460


                
                def update_ok_com():
                    roll_update_get=roll_entry_update_var.get()
                    class_update_get=class_entry_update_var.get()
                    english_update_get=english_entry_update_var.get()
                    science_update_get=science_entry_update_var.get()
                    social_update_get=Social_entry_update_var.get()
                    computer_update_get=computer_entry_update_var.get()
                    t_h_update_get=t_h_entry_update_var.get()
                    t_h_update_entry1_get=t_h_entry1_update_var.get()
                    maths_update_get=maths_entry_update_var.get()
                    #exam_update_get
                    
                   
                    table_update=log_username_get+'_'+class_update_get+'th'+'_'+t_h_update_get
                    statement_sql=f'update {table_update} set english = %s ,{t_h_update_get} = %s,mathematics = %s,general_science = %s,social_science = %s,computer_science = %s where rollno=%s'
                    statement_val=(english_update_get,t_h_update_entry1_get,maths_update_get,science_update_get,social_update_get,computer_update_get,roll_update_get)
                    print(statement_sql)
                    print(statement_val)
                    
                    cursor.execute(statement_sql,statement_val)
                    mydb.commit()

                global ok_update_btn    
                ok_update_btn=Button(dash,text='OK',width=10,font=('Bold',20),command=update_ok_com)
                ok_update_btn.place(x=1000,y=570)

                
                pass
            manage_update=Button(dash,text='UPDATE',width=14,height=1,font=('Bold',30),command=update_data)
            manage_update.place(x=540,y=360)
            

            def delete_data():
                roll_update.destroy()
                roll_entry_update.destroy()
                class_update.destroy()
                class_entry_update.destroy()
                english_update.destroy()
                english_entry_update.destroy()
                t_h_entry_update.destroy()
                t_h_entry1_update.destroy()
                science_update.destroy()
                science_entry_update.destroy()
                Social_update.destroy()
                Social_entry_update.destroy()
                computer_update.destroy()
                computer_entry_update.destroy()
                exam_label_update.destroy()
                exam_update_combo.destroy()
                maths_update.destroy()
                maths_entry_update.destroy()
                ok_update_btn.destroy()

                
                roll_delete=Label(dash,text='Roll.No',width=10,font=("bold",20))
                roll_delete.place(x=540,y=460)
                roll_entry_delete_var=StringVar()
                roll_entry_delete=Entry(dash,textvariable=roll_entry_delete_var,width=10,font=("bold",20))
                roll_entry_delete.place(x=700,y=460)

                class_delete=Label(dash,text='Class',width=10,font=("bold",20))
                class_delete.place(x=540,y=500)
                class_entry_delete_var=StringVar()
                class_entry_delete=Entry(dash,textvariable=class_entry_delete_var,width=10,font=("bold",20))
                class_entry_delete.place(x=700,y=500)

                
                t_h_entry_delete_var=StringVar()
                t_h_entry_delete=ttk.Combobox(dash,textvariable=t_h_entry_delete_var,font=("bold",15),width=5)
                t_h_entry_delete['values']=('telugu','hindi')
                t_h_entry_delete.current(0)
                t_h_entry_delete.place(x=540,y=540)

                
                def delete_ok_com():
                    roll_delete_get=roll_entry_delete_var.get()
                    class_delete_get=class_entry_delete_var.get()
                    t_h_delete_get=t_h_entry_delete_var.get()
                    #print(mydb.is_connected())
                    #print(log_username_get)
                    table_delete=log_username_get+'_'+class_delete_get+'th'+'_'+t_h_delete_get
                    del_statement_sql=f'delete from {table_delete} where rollno = {roll_delete_get}'
                    #del_statement_val=()
                    #print(del_statement_sql,del_statement_val)
                    cursor.execute(del_statement_sql)
                    mydb.commit()
                    pass

                manage_delete_btn=Button(dash,text='OK',font=('Bold',20),command=delete_ok_com)
                manage_delete_btn.place(x=700,y=580)
                pass
            manage_delete=Button(dash,text='DELETE',width=14,height=1,font=('Bold',30),command=delete_data)
            manage_delete.place(x=1140,y=360)
            
            def settings():
                setting_win=Toplevel()
                setting_win.title("Settings")
                #username setting starts
                setting_username=Label(setting_win,text="USERNAME",font=("Algerian",20))
                setting_username.place(x=0,y=0)

                setting_username_get=Label(setting_win,text=log_username_get,font=("bold",20))
                setting_username_get.place(x=180,y=0)
                #username setting starts

                #email setting starts
                setting_email=Label(setting_win,text="EMAIL",font=("Algerian",20))
                setting_email.place(x=0,y=50)
                setting_email_get=Label(setting_win,text=log_email_get,font=("bold",20))
                setting_email_get.place(x=180,y=50)
                #email setting starts

                #password setting starts
                setting_pass=Label(setting_win,text="PASSWORD",font=("Algerian",20))
                setting_pass.place(x=0,y=150)
                setting_pass_get=Label(setting_win,text="*********",font=("bold",20))
                setting_pass_get.place(x=180,y=150)
                def show_pass():
                    cursor.execute("select * from users")
                    get_pass_set_db=cursor.fetchall()
                    for x in get_pass_set_db:
                        if x[0]==log_username_get:
                            setting_pass_show1=x[2]
                    setting_pass_get.configure(text=setting_pass_show1)
                setting_pass_show=Button(setting_win,text="Show",command=show_pass,width=10)
                setting_pass_show.place(x=180,y=190)


                def change_pass():
                    setting_email_edit.configure(state='disabled')
                    change_pass_var_1=StringVar()
                    setting_change_pass_entry_1=Entry(setting_win,textvariable=change_pass_var_1,width=18,font=("bold",20))
                    setting_change_pass_entry_1.place(x=180,y=240)
                    change_pass_var_2=StringVar()
                    setting_change_pass_entry_2=Entry(setting_win,textvariable=change_pass_var_2,width=18,font=("bold",20))
                    setting_change_pass_entry_2.place(x=180,y=290)

                    def pass_change_ok():
                        setting_email_edit.configure(state='active')
                        change_pass_update_1=change_pass_var_1.get()
                        change_pass_update_2=change_pass_var_2.get()
                        if change_pass_update_1==change_pass_update_2:
                            change_sql_pass="update users set password = %s where username = %s"
                            change_val_pass=(change_pass_update_1,log_username_get)
                            cursor.execute(change_sql_pass,change_val_pass)
                            mydb.commit()
                        setting_pass_change.configure(state='active')
                        setting_change_pass_entry_1.destroy()
                        setting_change_pass_entry_2.destroy()
                        pass_change_ok_button.destroy()
                           
                    pass_change_ok_button=Button(setting_win,text="Ok",command=pass_change_ok,width=5,height=2)
                    pass_change_ok_button.place(x=460,y=288)
                    setting_pass_change.configure(state='disabled')
                setting_pass_change=Button(setting_win,text="Change",command=change_pass,width=10)
                setting_pass_change.place(x=300,y=190)
                #password setting ends

                #email edit button starts
                def edit_mail():
                    setting_pass_change.configure(state='disabled')
                    setting_pass.place(x=0,y=200)
                    setting_pass_get.place(x=180,y=200)
                    #global email_edit_var
                    email_edit_var=StringVar()
                    email_edit_entry=Entry(setting_win,textvariable=email_edit_var,width=18,font=("bold",20))
                    email_edit_entry.place(x=180,y=150)
                    setting_email_edit.configure(state='disabled')
                    setting_pass_show.place(x=180,y=240)
                    setting_pass_change.place(x=300,y=240)
                    
                    def email_edit_ok():
                        setting_pass_change.configure(state='active')
                        setting_pass_show.place(x=180,y=190)
                        setting_pass_change.place(x=300,y=190)
                        setting_pass.place(x=0,y=150)
                        setting_pass_get.place(x=180,y=150)
                        update_email=email_edit_var.get()
                        #print(update_email)
                        email_sql="update users set email = %s where username = %s"
                        email_val=(update_email,log_username_get)
                        cursor.execute(email_sql,email_val)
                        mydb.commit()
                        setting_email_get.configure(text=update_email)
                        
                        setting_email_edit.configure(state='active')
                        email_edit_entry.destroy()
                        email_edit_ok_button.destroy()
                    email_edit_ok_button=Button(setting_win,text="Ok",command=email_edit_ok,width=5,height=2)
                    email_edit_ok_button.place(x=460,y=148)
            
                setting_email_edit=Button(setting_win,text="Edit",command=edit_mail,width=10)
                setting_email_edit.place(x=180,y=100)
                #email edit button ends


                
                
                setting_win.geometry("550x700")
                setting_win.mainloop()
                
            profile_button=Button(dash,text="Profile",command=settings,width=15,font=("bold",20))
            profile_button.place(x=1250,y=60)

            #student data STARTS
            def _6th_student():            
                class_clicked='VI'
                
                student_window_6th = Toplevel()
                student_window_6th.title('STUDENT DATA ENTRY')
                #frame label
                frame_label_6th=Label(student_window_6th,text='',width=212,relief='groove',height=11,bd=20)
                frame_label_6th.place(x=0,y=0)

                frame_label_data_6th=Label(student_window_6th,text='',width=212,relief='groove',height=24,bd=20)
                frame_label_data_6th.place(x=0,y=200)

                frame_label_export_6th=Label(student_window_6th,text='',width=212,relief='groove',height=11,bd=20)
                frame_label_export_6th.place(x=0,y=600)
                #
                
                open_image_6th=Image.open(photo_fetch_address)
                dash_photo_6th=ImageTk.PhotoImage(open_image)
                photo_label_6th=Label(student_window_6th,image=dash_photo_6th)
                photo_label_6th.place(x=20,y=20)


                dash_school_label_6th=Label(student_window_6th,text=school_fetch_name,font=("Algerian",40))
                dash_school_label_6th.place(x=150,y=60)
                #####database entry starts


                #roll no of student starts
                rollno_label_6th=Label(student_window_6th,text='Roll No',font=("bold",20))
                rollno_label_6th.place(x=20,y=230)
                
                rollno_label_6th_var=IntVar()
                rollno_label_6th_entry=Entry(student_window_6th,textvariable=rollno_label_6th_var,width=20,font=("bold",20))
                rollno_label_6th_entry.place(x=180,y=230)
                #roll no of student ends


                #name of student starts
                name_label_6th=Label(student_window_6th,text='Name',font=("bold",20))
                name_label_6th.place(x=20,y=300)#+70
                
                name_label_6th_var=StringVar()
                name_label_6th_entry=Entry(student_window_6th,textvariable=name_label_6th_var,width=20,font=("bold",20))
                name_label_6th_entry.place(x=180,y=300)
                #name of student ends

                #section of student starts
                section_label_6th=Label(student_window_6th,text='Section',font=("bold",20))
                section_label_6th.place(x=600,y=230)#+70
                
                section_label_6th_var=StringVar()
                section_label_6th_combo=ttk.Combobox(student_window_6th,textvariable=section_label_6th_var,width=20,font=("bold",20))
                section_label_6th_combo['values']=('A','B','C')
                section_label_6th_combo.current(0)
                section_label_6th_combo.place(x=760,y=230)
                #section of student ends


                #exam of student starts
                exam_label_6th=Label(student_window_6th,text='Examination',font=("bold",20))
                exam_label_6th.place(x=600,y=300)#+70
                
                exam_label_6th_var=StringVar()
                exam_label_6th_combo=ttk.Combobox(student_window_6th,textvariable=exam_label_6th_var,width=20,font=("bold",20))
                exam_label_6th_combo['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                exam_label_6th_combo.current(0)
                exam_label_6th_combo.place(x=760,y=300)
                #exam of student ends
                

                ###subjects start
                subject_6th=Label(student_window_6th,text='Subject',font=("bold",20),foreground='#008f00')
                subject_6th.place(x=20,y=370)

                english_6th=Label(student_window_6th,text='English',font=("bold",20))
                english_6th.place(x=180,y=370)

                t_h_var_6th=StringVar()
                t_h_6th=ttk.Combobox(student_window_6th,textvariable=t_h_var_6th,font=("bold",20),width=6)
                t_h_6th['values']=('Telugu','Hindi')
                t_h_6th.current(0)
                t_h_6th.place(x=340,y=370)

                maths_6th=Label(student_window_6th,text='Mathematics',font=("bold",20))
                maths_6th.place(x=500,y=370)

                science_6th=Label(student_window_6th,text='General Science',font=("bold",20))
                science_6th.place(x=730,y=370)

                social_6th=Label(student_window_6th,text='Social Science',font=("bold",20))
                social_6th.place(x=1020,y=370)

                computer_6th=Label(student_window_6th,text='Computer Science',font=("bold",20))
                computer_6th.place(x=1270,y=370)

                ###subjects ends


                ###marks start
                marks_6th=Label(student_window_6th,text='Marks',font=("bold",20),foreground='#008f00')
                marks_6th.place(x=20,y=440)

                english_marks_var_6th=IntVar()
                english_marks_6th=Entry(student_window_6th,textvariable=english_marks_var_6th,font=("bold",20),width=6)
                english_marks_6th.place(x=180,y=440)

                t_h_marks_var_6th=IntVar()
                t_h_marks_6th=Entry(student_window_6th,textvariable=t_h_marks_var_6th,font=("bold",20),width=7)
                t_h_marks_6th.place(x=340,y=440)

                maths_marks_var_6th=IntVar()
                maths_marks_6th=Entry(student_window_6th,textvariable=maths_marks_var_6th,font=("bold",20),width=10)
                maths_marks_6th.place(x=500,y=440)

                science_marks_var_6th=IntVar()
                science_marks_6th=Entry(student_window_6th,textvariable=science_marks_var_6th,font=("bold",20),width=13)
                science_marks_6th.place(x=730,y=440)

                social_marks_var_6th=IntVar()
                social_marks_6th=Entry(student_window_6th,textvariable=social_marks_var_6th,font=("bold",20),width=12)
                social_marks_6th.place(x=1020,y=440)

                computer_marks_var_6th=IntVar()
                computer_marks_6th=Entry(student_window_6th,textvariable=computer_marks_var_6th,font=("bold",20),width=15)
                computer_marks_6th.place(x=1270,y=440)
                ###marks ends
                export_button_entry_var_6th=StringVar()
                export_button_entry_6th=Entry(student_window_6th,textvariable=export_button_entry_var_6th,font=("bold",20),width=15)
                export_button_entry_6th.place(x=180,y=710)


                
                
                #ok button db connecter starts
                def upload_marks_data_mysql():
                    #get values of details
                    main_data_name=name_label_6th_var.get()
                    main_data_rollno=rollno_label_6th_var.get()
                    main_data_section=section_label_6th_var.get()
                    main_data_exam=exam_label_6th_var.get()
                    local_class='VI'
                    #get values of marks
                    main_data_english=english_marks_var_6th.get()
                    main_data_t_h=t_h_marks_var_6th.get()
                    main_data_maths=maths_marks_var_6th.get()
                    main_data_science=science_marks_var_6th.get()
                    main_data_social=social_marks_var_6th.get()
                    main_data_computer=computer_marks_var_6th.get()
                    #hindi/telugu check var
                    check_t_h=t_h_var_6th.get()
                    
                    table_name_t_6th=log_username_get + '_6th_telugu'
                    table_name_h_6th=log_username_get + '_6th_hindi'
                    
                    if main_data_name and main_data_rollno and main_data_section and main_data_exam and main_data_english and main_data_t_h and main_data_maths and main_data_science and main_data_social and main_data_computer != None:
                        if check_t_h == 'Telugu':
                            query_t_6th=f'create table if not exists {table_name_t_6th}(rollno int(3) primary key,name varchar(50) not null,class varchar(4) not null,section char(1) not null,examination varchar(30) not null,english varchar(3) not null,telugu varchar(3) not null,mathematics varchar(3) not null,general_science varchar(3) not null,social_science varchar(3) not null,computer_science varchar(3) not null)'
                            cursor.execute(query_t_6th)
                            update_t_6th=f'insert into {table_name_t_6th} values({main_data_rollno},"{main_data_name}","{local_class}","{main_data_section}","{main_data_exam}",{main_data_english},{main_data_t_h},{main_data_maths},{main_data_science},{main_data_social},{main_data_computer})'
                            cursor.execute(update_t_6th)
                            mydb.commit()
                        elif check_t_h == 'Hindi':
                            query_h_6th=f'create table if not exists {table_name_h_6th}(rollno int(3) primary key,name varchar(30) not null,class varchar(4) not null,section char(1) not null,examination varchar(30) not null,english varchar(3) not null,hindi varchar(3) not null,mathematics varchar(3) not null,general_science varchar(3) not null,social_science varchar(3) not null,computer_science varchar(3) not null)'
                            cursor.execute(query_h_6th)
                            update_h_6th=f'insert into {table_name_h_6th} values({main_data_rollno},"{main_data_name}","{local_class}","{main_data_section}","{main_data_exam}",{main_data_english},{main_data_t_h},{main_data_maths},{main_data_science},{main_data_social},{main_data_computer})'
                            cursor.execute(update_h_6th)
                            mydb.commit()
                    else:
                            msg.showwarning("Warning","please fill all details")
                            
                    pass
                upload_marks_data=Button(student_window_6th,text='OK',font=("bold",20),command=upload_marks_data_mysql)
                upload_marks_data.place(x=1400,y=510)
                #ok button db connecter ends
                #####database entry ends

                #####export data starts 6th
                export_label_6th=Label(student_window_6th,text='EXPORT',foreground='red',font=("bold",20))
                export_label_6th.place(x=20,y=610)

                #class selection starts
                export_class_label_6th=Label(student_window_6th,text='Class',font=("bold",20))
                export_class_label_6th.place(x=20,y=650)

                export_class_var_6th=StringVar()
                export_class_entry_6th=ttk.Combobox(student_window_6th,textvariable=export_class_var_6th,font=("bold",20),width=6)
                export_class_entry_6th.place(x=160,y=650)
                export_class_entry_6th['values']=('6th','7th','8th','9th','10th','11th','12th')
                #class selection ends


                #section selection starts
                export_section_label_6th=Label(student_window_6th,text='Section',font=("bold",20))
                export_section_label_6th.place(x=300,y=650)

                export_section_var_6th=StringVar()
                export_section_entry_6th=ttk.Combobox(student_window_6th,textvariable=export_section_var_6th,font=("bold",20),width=6)
                export_section_entry_6th.place(x=440,y=650)
                export_section_entry_6th['values']=('A','B','C')
                #section selection ends

                #optional_sub selection starts
                export_optional_sub_label_6th=Label(student_window_6th,text='Optional Subject',font=("bold",20))
                export_optional_sub_label_6th.place(x=580,y=650)

                export_optional_sub_var_6th=StringVar()
                export_optional_sub_entry_6th=ttk.Combobox(student_window_6th,textvariable=export_optional_sub_var_6th,font=("bold",20),width=6)
                export_optional_sub_entry_6th.place(x=820,y=650)
                export_optional_sub_entry_6th['values']=('telugu','hindi')
                #optional_sub selection ends

                #exam selection starts
                export_exam_label_6th=Label(student_window_6th,text='Examination',font=("bold",20))
                export_exam_label_6th.place(x=1000,y=650)
                
                export_exam_var_6th=StringVar()
                export_exam_entry_6th=ttk.Combobox(student_window_6th,textvariable=export_exam_var_6th,font=("bold",20),width=15)
                export_exam_entry_6th.place(x=1190,y=650)
                export_exam_entry_6th['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                #exam selection ends


                #export button starts
                def export_6th():
                    exp_cls_6th=export_class_var_6th.get()
                    exp_sec_6th=export_section_var_6th.get()
                    exp_opt_sub_6th=export_optional_sub_var_6th.get()
                    exp_exami_6th=export_exam_var_6th.get()
                    export_name_6th_get=export_button_entry_var_6th.get()
                    print(exp_exami_6th)
                    error_fix_6th='error'
                    if exp_cls_6th and exp_sec_6th and exp_opt_sub_6th  and exp_exami_6th and export_name_6th_get and error_fix_6th != None:
                        
                        class_list_6th=['6th','7th','8th','9th','10th','11th','12th']
                        for i_6th in class_list_6th:
                            if exp_cls_6th == i_6th:
                                opt_sub_tem_6th= f"{exp_opt_sub_6th}"
                                print(opt_sub_tem_6th)
                                temp_cls_tab_6th=log_username_get + '_' + i_6th + '_' + str(opt_sub_tem_6th)
                                cursor.execute(f'select * from {temp_cls_tab_6th} where section = "{exp_sec_6th}" and examination = "{exp_exami_6th}"')
                                get_exp_6th=cursor.fetchall()
                                
                                print(get_exp_6th)
                                wb_6th=Workbook()
                                sheet_6th=wb_6th.active
                                half_6th=["rollno","name","class","section","examination","english"]
                                half_6th.append(opt_sub_tem_6th)
                                full_6th=["mathematics","general_science","social_science","computer_science"]
                                complete_list_6th=half_6th + full_6th
                                print(type(complete_list_6th))
                                sheet_6th.append(complete_list_6th)
        
                                
                                for ex_i_6th in get_exp_6th:
                                    sheet_6th.append(ex_i_6th)
                                export_name_6th_get=export_button_entry_var_6th.get()
                                export_filename_get=filedialog.askdirectory()
                                savespot_6th=str(export_filename_get)

                                wb_6th.save(savespot_6th + '/' + str(export_name_6th_get) + '.xlsx')
                                print(savespot_6th + str(export_name_6th_get) + '.xlsx')


                    else:
                        msg.showwarning('Warning','Please fill all details')
                    pass
                export_button_6th=Button(student_window_6th,text='Save AS',command=export_6th,font=('bold',20))
                export_button_6th.place(x=20,y=700)
                #export button ends


                
                #####export data ends 6th
                student_window_6th.geometry('1920x1080')
                student_window_6th.mainloop()
            #student data ENDS
            dash_6th=Button(dash,text='VI',font=('bold',30),width=20,command=_6th_student)
            dash_6th.place(x=20,y=210)

            def _7th_student():            
                student_window_7th = Toplevel()
                student_window_7th.title('STUDENT DATA ENTRY')
                #frame label
                frame_label_7th=Label(student_window_7th,text='',width=212,relief='groove',height=11,bd=20)
                frame_label_7th.place(x=0,y=0)

                frame_label_data_7th=Label(student_window_7th,text='',width=212,relief='groove',height=24,bd=20)
                frame_label_data_7th.place(x=0,y=200)

                frame_label_export_7th=Label(student_window_7th,text='',width=212,relief='groove',height=11,bd=20)
                frame_label_export_7th.place(x=0,y=600)
                #
                
                open_image_7th=Image.open(photo_fetch_address)
                dash_photo_7th=ImageTk.PhotoImage(open_image)
                photo_label_7th=Label(student_window_7th,image=dash_photo_7th)
                photo_label_7th.place(x=20,y=20)


                dash_school_label_7th=Label(student_window_7th,text=school_fetch_name,font=("Algerian",40))
                dash_school_label_7th.place(x=150,y=60)
                #####database entry starts


                #roll no of student starts
                rollno_label_7th=Label(student_window_7th,text='Roll No',font=("bold",20))
                rollno_label_7th.place(x=20,y=230)
                
                rollno_label_7th_var=IntVar()
                rollno_label_7th_entry=Entry(student_window_7th,textvariable=rollno_label_7th_var,width=20,font=("bold",20))
                rollno_label_7th_entry.place(x=180,y=230)
                #roll no of student ends


                #name of student starts
                name_label_7th=Label(student_window_7th,text='Name',font=("bold",20))
                name_label_7th.place(x=20,y=300)#+70
                
                name_label_7th_var=StringVar()
                name_label_7th_entry=Entry(student_window_7th,textvariable=name_label_7th_var,width=20,font=("bold",20))
                name_label_7th_entry.place(x=180,y=300)
                #name of student ends

                #section of student starts
                section_label_7th=Label(student_window_7th,text='Section',font=("bold",20))
                section_label_7th.place(x=600,y=230)#+70
                
                section_label_7th_var=StringVar()
                section_label_7th_combo=ttk.Combobox(student_window_7th,textvariable=section_label_7th_var,width=20,font=("bold",20))
                section_label_7th_combo['values']=('A','B','C')
                section_label_7th_combo.current(0)
                section_label_7th_combo.place(x=760,y=230)
                #section of student ends


                #exam of student starts
                exam_label_7th=Label(student_window_7th,text='Examination',font=("bold",20))
                exam_label_7th.place(x=600,y=300)#+70
                
                exam_label_7th_var=StringVar()
                exam_label_7th_combo=ttk.Combobox(student_window_7th,textvariable=exam_label_7th_var,width=20,font=("bold",20))
                exam_label_7th_combo['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                exam_label_7th_combo.current(0)
                exam_label_7th_combo.place(x=760,y=300)
                #exam of student ends
                

                ###subjects start
                subject_7th=Label(student_window_7th,text='Subject',font=("bold",20),foreground='#008f00')
                subject_7th.place(x=20,y=370)

                english_7th=Label(student_window_7th,text='English',font=("bold",20))
                english_7th.place(x=180,y=370)

                t_h_var_7th=StringVar()
                t_h_7th=ttk.Combobox(student_window_7th,textvariable=t_h_var_7th,font=("bold",20),width=6)
                t_h_7th['values']=('Telugu','Hindi')
                t_h_7th.current(0)
                t_h_7th.place(x=340,y=370)

                maths_7th=Label(student_window_7th,text='Mathematics',font=("bold",20))
                maths_7th.place(x=500,y=370)

                science_7th=Label(student_window_7th,text='General Science',font=("bold",20))
                science_7th.place(x=730,y=370)

                social_7th=Label(student_window_7th,text='Social Science',font=("bold",20))
                social_7th.place(x=1020,y=370)

                computer_7th=Label(student_window_7th,text='Computer Science',font=("bold",20))
                computer_7th.place(x=1270,y=370)

                ###subjects ends


                ###marks start
                marks_7th=Label(student_window_7th,text='Marks',font=("bold",20),foreground='#008f00')
                marks_7th.place(x=20,y=440)

                english_marks_var_7th=IntVar()
                english_marks_7th=Entry(student_window_7th,textvariable=english_marks_var_7th,font=("bold",20),width=6)
                english_marks_7th.place(x=180,y=440)

                t_h_marks_var_7th=IntVar()
                t_h_marks_7th=Entry(student_window_7th,textvariable=t_h_marks_var_7th,font=("bold",20),width=7)
                t_h_marks_7th.place(x=340,y=440)

                maths_marks_var_7th=IntVar()
                maths_marks_7th=Entry(student_window_7th,textvariable=maths_marks_var_7th,font=("bold",20),width=10)
                maths_marks_7th.place(x=500,y=440)

                science_marks_var_7th=IntVar()
                science_marks_7th=Entry(student_window_7th,textvariable=science_marks_var_7th,font=("bold",20),width=13)
                science_marks_7th.place(x=730,y=440)

                social_marks_var_7th=IntVar()
                social_marks_7th=Entry(student_window_7th,textvariable=social_marks_var_7th,font=("bold",20),width=12)
                social_marks_7th.place(x=1020,y=440)

                computer_marks_var_7th=IntVar()
                computer_marks_7th=Entry(student_window_7th,textvariable=computer_marks_var_7th,font=("bold",20),width=15)
                computer_marks_7th.place(x=1270,y=440)
                ###marks ends
                export_button_entry_var_7th=StringVar()
                export_button_entry_7th=Entry(student_window_7th,textvariable=export_button_entry_var_7th,font=("bold",20),width=15)
                export_button_entry_7th.place(x=180,y=710)


                
                
                #ok button db connecter starts
                def upload_marks_data_mysql_7th():
                    #get values of details
                    main_data_name_7th=name_label_7th_var.get()
                    main_data_rollno_7th=rollno_label_7th_var.get()
                    main_data_section_7th=section_label_7th_var.get()
                    main_data_exam_7th=exam_label_7th_var.get()
                    local_class_7th='VII'
                    #get values of marks
                    main_data_english_7th=english_marks_var_7th.get()
                    main_data_t_h_7th=t_h_marks_var_7th.get()
                    main_data_maths_7th=maths_marks_var_7th.get()
                    main_data_science_7th=science_marks_var_7th.get()
                    main_data_social_7th=social_marks_var_7th.get()
                    main_data_computer_7th=computer_marks_var_7th.get()
                    #hindi/telugu check var
                    check_t_h_7th=t_h_var_7th.get()
                    
                    table_name_t_7th=log_username_get + '_7th_telugu'
                    table_name_h_7th=log_username_get + '_7th_hindi'
                    
                    if main_data_name_7th and main_data_rollno_7th and main_data_section_7th and main_data_exam_7th and main_data_english_7th and main_data_t_h_7th and main_data_maths_7th and main_data_science_7th and main_data_social_7th and main_data_computer_7th != None:
                        if check_t_h_7th == 'Telugu':
                            query_t_7th=f'create table if not exists {table_name_t_7th}(rollno int(3) primary key,name varchar(50) not null,class varchar(4) not null,section char(1) not null,examination varchar(30) not null,english varchar(3) not null,telugu varchar(3) not null,mathematics varchar(3) not null,general_science varchar(3) not null,social_science varchar(3) not null,computer_science varchar(3) not null)'
                            cursor.execute(query_t_7th)
                            update_t_7th=f'insert into {table_name_t_7th} values({main_data_rollno_7th},"{main_data_name_7th}","{local_class_7th}","{main_data_section_7th}","{main_data_exam_7th}",{main_data_english_7th},{main_data_t_h_7th},{main_data_maths_7th},{main_data_science_7th},{main_data_social_7th},{main_data_computer_7th})'
                            cursor.execute(update_t_7th)
                            mydb.commit()
                        elif check_t_h_7th == 'Hindi':
                            query_h_7th=f'create table if not exists {table_name_h_7th}(rollno int(3) primary key,name varchar(30) not null,class varchar(4) not null,section char(1) not null,examination varchar(30) not null,english varchar(3) not null,hindi varchar(3) not null,mathematics varchar(3) not null,general_science varchar(3) not null,social_science varchar(3) not null,computer_science varchar(3) not null)'
                            cursor.execute(query_h_7th)
                            update_h_7th=f'insert into {table_name_h_7th} values({main_data_rollno_7th},"{main_data_name_7th}","{local_class_7th}","{main_data_section_7th}","{main_data_exam_7th}",{main_data_english_7th},{main_data_t_h_7th},{main_data_maths_7th},{main_data_science_7th},{main_data_social_7th},{main_data_computer_7th})'
                            cursor.execute(update_h_7th)
                            mydb.commit()
                    else:
                            msg.showwarning("Warning","please fill all details")
                            
                    pass
                upload_marks_data_7th=Button(student_window_7th,text='OK',font=("bold",20),command=upload_marks_data_mysql_7th)
                upload_marks_data_7th.place(x=1400,y=510)
                #ok button db connecter ends
                #####database entry ends

                #####export data starts 7th
                export_label_7th=Label(student_window_7th,text='EXPORT',foreground='red',font=("bold",20))
                export_label_7th.place(x=20,y=610)

                #class selection starts
                export_class_label_7th=Label(student_window_7th,text='Class',font=("bold",20))
                export_class_label_7th.place(x=20,y=650)

                export_class_var_7th=StringVar()
                export_class_entry_7th=ttk.Combobox(student_window_7th,textvariable=export_class_var_7th,font=("bold",20),width=6)
                export_class_entry_7th.place(x=160,y=650)
                export_class_entry_7th['values']=('6th','7th','8th','9th','10th','11th','12th')
                #class selection ends


                #section selection starts
                export_section_label_7th=Label(student_window_7th,text='Section',font=("bold",20))
                export_section_label_7th.place(x=300,y=650)

                export_section_var_7th=StringVar()
                export_section_entry_7th=ttk.Combobox(student_window_7th,textvariable=export_section_var_7th,font=("bold",20),width=6)
                export_section_entry_7th.place(x=440,y=650)
                export_section_entry_7th['values']=('A','B','C')
                #section selection ends

                #optional_sub selection starts
                export_optional_sub_label_7th=Label(student_window_7th,text='Optional Subject',font=("bold",20))
                export_optional_sub_label_7th.place(x=580,y=650)

                export_optional_sub_var_7th=StringVar()
                export_optional_sub_entry_7th=ttk.Combobox(student_window_7th,textvariable=export_optional_sub_var_7th,font=("bold",20),width=6)
                export_optional_sub_entry_7th.place(x=820,y=650)
                export_optional_sub_entry_7th['values']=('telugu','hindi')
                #optional_sub selection ends

                #exam selection starts
                export_exam_label_7th=Label(student_window_7th,text='Examination',font=("bold",20))
                export_exam_label_7th.place(x=1000,y=650)
                
                export_exam_var_7th=StringVar()
                export_exam_entry_7th=ttk.Combobox(student_window_7th,textvariable=export_exam_var_7th,font=("bold",20),width=15)
                export_exam_entry_7th.place(x=1190,y=650)
                export_exam_entry_7th['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                #exam selection ends


                #export button starts
                def export_7th():
                    exp_cls_7th=export_class_var_7th.get()
                    exp_sec_7th=export_section_var_7th.get()
                    exp_opt_sub_7th=export_optional_sub_var_7th.get()
                    exp_exami_7th=export_exam_var_7th.get()
                    export_name_7th_get=export_button_entry_var_7th.get()
                    print(exp_exami_7th)
                    error_fix_7th='error'
                    if exp_cls_7th and exp_sec_7th and exp_opt_sub_7th  and exp_exami_7th and export_name_7th_get and error_fix_7th != None:
                        
                        class_list_7th=['6th','7th','8th','9th','10th','11th','12th']
                        for i_7th in class_list_7th:
                            if exp_cls_7th == i_7th:
                                opt_sub_tem_7th= f"{exp_opt_sub_7th}"
                                print(opt_sub_tem_7th)
                                temp_cls_tab_7th=log_username_get + '_' + i_7th + '_' + str(opt_sub_tem_7th)
                                cursor.execute(f'select * from {temp_cls_tab_7th} where section = "{exp_sec_7th}" and examination = "{exp_exami_7th}"')
                                get_exp_7th=cursor.fetchall()
                                
                                print(get_exp_7th)
                                wb_7th=Workbook()
                                sheet_7th=wb_7th.active
                                half_7th=["rollno","name","class","section","examination","english"]
                                half_7th.append(opt_sub_tem_7th)
                                full_7th=["mathematics","general_science","social_science","computer_science"]
                                complete_list_7th=half_7th + full_7th
                                print(type(complete_list_7th))
                                sheet_7th.append(complete_list_7th)
        
                                
                                for ex_i_7th in get_exp_7th:
                                    sheet_7th.append(ex_i_7th)
                                export_name_7th_get=export_button_entry_var_7th.get()
                                export_filename_get_7th=filedialog.askdirectory()
                                savespot_7th=str(export_filename_get_7th)

                                wb_7th.save(savespot_7th + '/' + str(export_name_7th_get) + '.xlsx')
                                print(savespot_7th + str(export_name_7th_get) + '.xlsx')


                    else:
                        msg.showwarning('Warning','Please fill all details')
                    pass
                export_button_7th=Button(student_window_7th,text='Save AS',command=export_7th,font=('bold',20))
                export_button_7th.place(x=20,y=700)
                #export button ends


                
                #####export data ends 7th
                student_window_7th.geometry('1920x1080')
                student_window_7th.mainloop()
            #student data ENDS

             
            dash_7th=Button(dash,text='VII',font=('bold',30),width=20,command=_7th_student)
            dash_7th.place(x=20,y=290)


            def _8th_student():            
                student_window_8th = Toplevel()
                student_window_8th.title('STUDENT DATA ENTRY')
                #frame label
                frame_label_8th=Label(student_window_8th,text='',width=212,relief='groove',height=11,bd=20)
                frame_label_8th.place(x=0,y=0)

                frame_label_data_8th=Label(student_window_8th,text='',width=212,relief='groove',height=24,bd=20)
                frame_label_data_8th.place(x=0,y=200)

                frame_label_export_8th=Label(student_window_8th,text='',width=212,relief='groove',height=11,bd=20)
                frame_label_export_8th.place(x=0,y=600)
                #
                
                open_image_8th=Image.open(photo_fetch_address)
                dash_photo_8th=ImageTk.PhotoImage(open_image)
                photo_label_8th=Label(student_window_8th,image=dash_photo_8th)
                photo_label_8th.place(x=20,y=20)


                dash_school_label_8th=Label(student_window_8th,text=school_fetch_name,font=("Algerian",40))
                dash_school_label_8th.place(x=150,y=60)
                #####database entry starts


                #roll no of student starts
                rollno_label_8th=Label(student_window_8th,text='Roll No',font=("bold",20))
                rollno_label_8th.place(x=20,y=230)
                
                rollno_label_8th_var=IntVar()
                rollno_label_8th_entry=Entry(student_window_8th,textvariable=rollno_label_8th_var,width=20,font=("bold",20))
                rollno_label_8th_entry.place(x=180,y=230)
                #roll no of student ends


                #name of student starts
                name_label_8th=Label(student_window_8th,text='Name',font=("bold",20))
                name_label_8th.place(x=20,y=300)#+70
                
                name_label_8th_var=StringVar()
                name_label_8th_entry=Entry(student_window_8th,textvariable=name_label_8th_var,width=20,font=("bold",20))
                name_label_8th_entry.place(x=180,y=300)
                #name of student ends

                #section of student starts
                section_label_8th=Label(student_window_8th,text='Section',font=("bold",20))
                section_label_8th.place(x=600,y=230)#+70
                
                section_label_8th_var=StringVar()
                section_label_8th_combo=ttk.Combobox(student_window_8th,textvariable=section_label_8th_var,width=20,font=("bold",20))
                section_label_8th_combo['values']=('A','B','C')
                section_label_8th_combo.current(0)
                section_label_8th_combo.place(x=760,y=230)
                #section of student ends


                #exam of student starts
                exam_label_8th=Label(student_window_8th,text='Examination',font=("bold",20))
                exam_label_8th.place(x=600,y=300)#+70
                
                exam_label_8th_var=StringVar()
                exam_label_8th_combo=ttk.Combobox(student_window_8th,textvariable=exam_label_8th_var,width=20,font=("bold",20))
                exam_label_8th_combo['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                exam_label_8th_combo.current(0)
                exam_label_8th_combo.place(x=760,y=300)
                #exam of student ends
                

                ###subjects start
                subject_8th=Label(student_window_8th,text='Subject',font=("bold",20),foreground='#008f00')
                subject_8th.place(x=20,y=370)

                english_8th=Label(student_window_8th,text='English',font=("bold",20))
                english_8th.place(x=180,y=370)

                t_h_var_8th=StringVar()
                t_h_8th=ttk.Combobox(student_window_8th,textvariable=t_h_var_8th,font=("bold",20),width=6)
                t_h_8th['values']=('Telugu','Hindi')
                t_h_8th.current(0)
                t_h_8th.place(x=340,y=370)

                maths_8th=Label(student_window_8th,text='Mathematics',font=("bold",20))
                maths_8th.place(x=500,y=370)

                science_8th=Label(student_window_8th,text='General Science',font=("bold",20))
                science_8th.place(x=730,y=370)

                social_8th=Label(student_window_8th,text='Social Science',font=("bold",20))
                social_8th.place(x=1020,y=370)

                computer_8th=Label(student_window_8th,text='Computer Science',font=("bold",20))
                computer_8th.place(x=1270,y=370)

                ###subjects ends


                ###marks start
                marks_8th=Label(student_window_8th,text='Marks',font=("bold",20),foreground='#008f00')
                marks_8th.place(x=20,y=440)

                english_marks_var_8th=IntVar()
                english_marks_8th=Entry(student_window_8th,textvariable=english_marks_var_8th,font=("bold",20),width=6)
                english_marks_8th.place(x=180,y=440)

                t_h_marks_var_8th=IntVar()
                t_h_marks_8th=Entry(student_window_8th,textvariable=t_h_marks_var_8th,font=("bold",20),width=7)
                t_h_marks_8th.place(x=340,y=440)

                maths_marks_var_8th=IntVar()
                maths_marks_8th=Entry(student_window_8th,textvariable=maths_marks_var_8th,font=("bold",20),width=10)
                maths_marks_8th.place(x=500,y=440)

                science_marks_var_8th=IntVar()
                science_marks_8th=Entry(student_window_8th,textvariable=science_marks_var_8th,font=("bold",20),width=13)
                science_marks_8th.place(x=730,y=440)

                social_marks_var_8th=IntVar()
                social_marks_8th=Entry(student_window_8th,textvariable=social_marks_var_8th,font=("bold",20),width=12)
                social_marks_8th.place(x=1020,y=440)

                computer_marks_var_8th=IntVar()
                computer_marks_8th=Entry(student_window_8th,textvariable=computer_marks_var_8th,font=("bold",20),width=15)
                computer_marks_8th.place(x=1270,y=440)
                ###marks ends
                export_button_entry_var_8th=StringVar()
                export_button_entry_8th=Entry(student_window_8th,textvariable=export_button_entry_var_8th,font=("bold",20),width=15)
                export_button_entry_8th.place(x=180,y=710)


                
                
                #ok button db connecter starts
                def upload_marks_data_mysql_8th():
                    #get values of details
                    main_data_name_8th=name_label_8th_var.get()
                    main_data_rollno_8th=rollno_label_8th_var.get()
                    main_data_section_8th=section_label_8th_var.get()
                    main_data_exam_8th=exam_label_8th_var.get()
                    local_class_8th='VIII'
                    #get values of marks
                    main_data_english_8th=english_marks_var_8th.get()
                    main_data_t_h_8th=t_h_marks_var_8th.get()
                    main_data_maths_8th=maths_marks_var_8th.get()
                    main_data_science_8th=science_marks_var_8th.get()
                    main_data_social_8th=social_marks_var_8th.get()
                    main_data_computer_8th=computer_marks_var_8th.get()
                    #hindi/telugu check var
                    check_t_h_8th=t_h_var_8th.get()
                    
                    table_name_t_8th=log_username_get + '_8th_telugu'
                    table_name_h_8th=log_username_get + '_8th_hindi'
                    
                    if main_data_name_8th and main_data_rollno_8th and main_data_section_8th and main_data_exam_8th and main_data_english_8th and main_data_t_h_8th and main_data_maths_8th and main_data_science_8th and main_data_social_8th and main_data_computer_8th != None:
                        if check_t_h_8th == 'Telugu':
                            query_t_8th=f'create table if not exists {table_name_t_8th}(rollno int(3) primary key,name varchar(50) not null,class varchar(4) not null,section char(1) not null,examination varchar(30) not null,english varchar(3) not null,telugu varchar(3) not null,mathematics varchar(3) not null,general_science varchar(3) not null,social_science varchar(3) not null,computer_science varchar(3) not null)'
                            cursor.execute(query_t_8th)
                            update_t_8th=f'insert into {table_name_t_8th} values({main_data_rollno_8th},"{main_data_name_8th}","{local_class_8th}","{main_data_section_8th}","{main_data_exam_8th}",{main_data_english_8th},{main_data_t_h_8th},{main_data_maths_8th},{main_data_science_8th},{main_data_social_8th},{main_data_computer_8th})'
                            cursor.execute(update_t_8th)
                            mydb.commit()
                        elif check_t_h_8th == 'Hindi':
                            query_h_8th=f'create table if not exists {table_name_h_8th}(rollno int(3) primary key,name varchar(30) not null,class varchar(4) not null,section char(1) not null,examination varchar(30) not null,english varchar(3) not null,hindi varchar(3) not null,mathematics varchar(3) not null,general_science varchar(3) not null,social_science varchar(3) not null,computer_science varchar(3) not null)'
                            cursor.execute(query_h_8th)
                            update_h_8th=f'insert into {table_name_h_8th} values({main_data_rollno_8th},"{main_data_name_8th}","{local_class_8th}","{main_data_section_8th}","{main_data_exam_8th}",{main_data_english_8th},{main_data_t_h_8th},{main_data_maths_8th},{main_data_science_8th},{main_data_social_8th},{main_data_computer_8th})'
                            cursor.execute(update_h_8th)
                            mydb.commit()
                    else:
                            msg.showwarning("Warning","please fill all details")
                            
                    pass
                upload_marks_data_8th=Button(student_window_8th,text='OK',font=("bold",20),command=upload_marks_data_mysql_8th)
                upload_marks_data_8th.place(x=1400,y=510)
                #ok button db connecter ends
                #####database entry ends

                #####export data starts 8th
                export_label_8th=Label(student_window_8th,text='EXPORT',foreground='red',font=("bold",20))
                export_label_8th.place(x=20,y=610)

                #class selection starts
                export_class_label_8th=Label(student_window_8th,text='Class',font=("bold",20))
                export_class_label_8th.place(x=20,y=650)

                export_class_var_8th=StringVar()
                export_class_entry_8th=ttk.Combobox(student_window_8th,textvariable=export_class_var_8th,font=("bold",20),width=6)
                export_class_entry_8th.place(x=160,y=650)
                export_class_entry_8th['values']=('6th','7th','8th','9th','10th','11th','12th')
                #class selection ends


                #section selection starts
                export_section_label_8th=Label(student_window_8th,text='Section',font=("bold",20))
                export_section_label_8th.place(x=300,y=650)

                export_section_var_8th=StringVar()
                export_section_entry_8th=ttk.Combobox(student_window_8th,textvariable=export_section_var_8th,font=("bold",20),width=6)
                export_section_entry_8th.place(x=440,y=650)
                export_section_entry_8th['values']=('A','B','C')
                #section selection ends

                #optional_sub selection starts
                export_optional_sub_label_8th=Label(student_window_8th,text='Optional Subject',font=("bold",20))
                export_optional_sub_label_8th.place(x=580,y=650)

                export_optional_sub_var_8th=StringVar()
                export_optional_sub_entry_8th=ttk.Combobox(student_window_8th,textvariable=export_optional_sub_var_8th,font=("bold",20),width=6)
                export_optional_sub_entry_8th.place(x=820,y=650)
                export_optional_sub_entry_8th['values']=('telugu','hindi')
                #optional_sub selection ends

                #exam selection starts
                export_exam_label_8th=Label(student_window_8th,text='Examination',font=("bold",20))
                export_exam_label_8th.place(x=1000,y=650)
                
                export_exam_var_8th=StringVar()
                export_exam_entry_8th=ttk.Combobox(student_window_8th,textvariable=export_exam_var_8th,font=("bold",20),width=15)
                export_exam_entry_8th.place(x=1190,y=650)
                export_exam_entry_8th['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                #exam selection ends


                #export button starts
                def export_8th():
                    exp_cls_8th=export_class_var_8th.get()
                    exp_sec_8th=export_section_var_8th.get()
                    exp_opt_sub_8th=export_optional_sub_var_8th.get()
                    exp_exami_8th=export_exam_var_8th.get()
                    export_name_8th_get=export_button_entry_var_8th.get()
                    print(exp_exami_8th)
                    error_fix_8th='error'
                    if exp_cls_8th and exp_sec_8th and exp_opt_sub_8th  and exp_exami_8th and export_name_8th_get and error_fix_8th != None:
                        
                        class_list_8th=['6th','7th','8th','9th','10th','11th','12th']
                        for i_8th in class_list_8th:
                            if exp_cls_8th == i_8th:
                                opt_sub_tem_8th= f"{exp_opt_sub_8th}"
                                print(opt_sub_tem_8th)
                                temp_cls_tab_8th=log_username_get + '_' + i_8th + '_' + str(opt_sub_tem_8th)
                                cursor.execute(f'select * from {temp_cls_tab_8th} where section = "{exp_sec_8th}" and examination = "{exp_exami_8th}"')
                                get_exp_8th=cursor.fetchall()
                                
                                print(get_exp_8th)
                                wb_8th=Workbook()
                                sheet_8th=wb_8th.active
                                half_8th=["rollno","name","class","section","examination","english"]
                                half_8th.append(opt_sub_tem_8th)
                                full_8th=["mathematics","general_science","social_science","computer_science"]
                                complete_list_8th=half_8th + full_8th
                                print(type(complete_list_8th))
                                sheet_8th.append(complete_list_8th)
        
                                
                                for ex_i_8th in get_exp_8th:
                                    sheet_8th.append(ex_i_8th)
                                export_name_8th_get=export_button_entry_var_8th.get()
                                export_filename_get_8th=filedialog.askdirectory()
                                savespot_8th=str(export_filename_get_8th)

                                wb_8th.save(savespot_8th + '/' + str(export_name_8th_get) + '.xlsx')
                                print(savespot_8th + str(export_name_8th_get) + '.xlsx')


                    else:
                        msg.showwarning('Warning','Please fill all details')
                    pass
                export_button_8th=Button(student_window_8th,text='Save AS',command=export_8th,font=('bold',20))
                export_button_8th.place(x=20,y=700)
                #export button ends


                
                #####export data ends 8th
                student_window_8th.geometry('1920x1080')
                student_window_8th.mainloop()
            #student data ENDS

            dash_8th=Button(dash,text='VIII',font=('bold',30),width=20,command = _8th_student)
            dash_8th.place(x=20,y=370)


            def _9th_student():            
                student_window_9th = Toplevel()
                student_window_9th.title('STUDENT DATA ENTRY')
                #frame label
                frame_label_9th=Label(student_window_9th,text='',width=212,relief='groove',height=11,bd=20)
                frame_label_9th.place(x=0,y=0)

                frame_label_data_9th=Label(student_window_9th,text='',width=212,relief='groove',height=24,bd=20)
                frame_label_data_9th.place(x=0,y=200)

                frame_label_export_9th=Label(student_window_9th,text='',width=212,relief='groove',height=11,bd=20)
                frame_label_export_9th.place(x=0,y=600)
                #
                
                open_image_9th=Image.open(photo_fetch_address)
                dash_photo_9th=ImageTk.PhotoImage(open_image)
                photo_label_9th=Label(student_window_9th,image=dash_photo_9th)
                photo_label_9th.place(x=20,y=20)


                dash_school_label_9th=Label(student_window_9th,text=school_fetch_name,font=("Algerian",40))
                dash_school_label_9th.place(x=150,y=60)
                #####database entry starts


                #roll no of student starts
                rollno_label_9th=Label(student_window_9th,text='Roll No',font=("bold",20))
                rollno_label_9th.place(x=20,y=230)
                
                rollno_label_9th_var=IntVar()
                rollno_label_9th_entry=Entry(student_window_9th,textvariable=rollno_label_9th_var,width=20,font=("bold",20))
                rollno_label_9th_entry.place(x=180,y=230)
                #roll no of student ends


                #name of student starts
                name_label_9th=Label(student_window_9th,text='Name',font=("bold",20))
                name_label_9th.place(x=20,y=300)#+70
                
                name_label_9th_var=StringVar()
                name_label_9th_entry=Entry(student_window_9th,textvariable=name_label_9th_var,width=20,font=("bold",20))
                name_label_9th_entry.place(x=180,y=300)
                #name of student ends

                #section of student starts
                section_label_9th=Label(student_window_9th,text='Section',font=("bold",20))
                section_label_9th.place(x=600,y=230)#+70
                
                section_label_9th_var=StringVar()
                section_label_9th_combo=ttk.Combobox(student_window_9th,textvariable=section_label_9th_var,width=20,font=("bold",20))
                section_label_9th_combo['values']=('A','B','C')
                section_label_9th_combo.current(0)
                section_label_9th_combo.place(x=760,y=230)
                #section of student ends


                #exam of student starts
                exam_label_9th=Label(student_window_9th,text='Examination',font=("bold",20))
                exam_label_9th.place(x=600,y=300)#+70
                
                exam_label_9th_var=StringVar()
                exam_label_9th_combo=ttk.Combobox(student_window_9th,textvariable=exam_label_9th_var,width=20,font=("bold",20))
                exam_label_9th_combo['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                exam_label_9th_combo.current(0)
                exam_label_9th_combo.place(x=760,y=300)
                #exam of student ends
                

                ###subjects start
                subject_9th=Label(student_window_9th,text='Subject',font=("bold",20),foreground='#008f00')
                subject_9th.place(x=20,y=370)

                english_9th=Label(student_window_9th,text='English',font=("bold",20))
                english_9th.place(x=180,y=370)

                t_h_var_9th=StringVar()
                t_h_9th=ttk.Combobox(student_window_9th,textvariable=t_h_var_9th,font=("bold",20),width=6)
                t_h_9th['values']=('Telugu','Hindi')
                t_h_9th.current(0)
                t_h_9th.place(x=340,y=370)

                maths_9th=Label(student_window_9th,text='Mathematics',font=("bold",20))
                maths_9th.place(x=500,y=370)

                science_9th=Label(student_window_9th,text='General Science',font=("bold",20))
                science_9th.place(x=730,y=370)

                social_9th=Label(student_window_9th,text='Social Science',font=("bold",20))
                social_9th.place(x=1020,y=370)

                computer_9th=Label(student_window_9th,text='Computer Science',font=("bold",20))
                computer_9th.place(x=1270,y=370)

                ###subjects ends


                ###marks start
                marks_9th=Label(student_window_9th,text='Marks',font=("bold",20),foreground='#008f00')
                marks_9th.place(x=20,y=440)

                english_marks_var_9th=IntVar()
                english_marks_9th=Entry(student_window_9th,textvariable=english_marks_var_9th,font=("bold",20),width=6)
                english_marks_9th.place(x=180,y=440)

                t_h_marks_var_9th=IntVar()
                t_h_marks_9th=Entry(student_window_9th,textvariable=t_h_marks_var_9th,font=("bold",20),width=7)
                t_h_marks_9th.place(x=340,y=440)

                maths_marks_var_9th=IntVar()
                maths_marks_9th=Entry(student_window_9th,textvariable=maths_marks_var_9th,font=("bold",20),width=10)
                maths_marks_9th.place(x=500,y=440)

                science_marks_var_9th=IntVar()
                science_marks_9th=Entry(student_window_9th,textvariable=science_marks_var_9th,font=("bold",20),width=13)
                science_marks_9th.place(x=730,y=440)

                social_marks_var_9th=IntVar()
                social_marks_9th=Entry(student_window_9th,textvariable=social_marks_var_9th,font=("bold",20),width=12)
                social_marks_9th.place(x=1020,y=440)

                computer_marks_var_9th=IntVar()
                computer_marks_9th=Entry(student_window_9th,textvariable=computer_marks_var_9th,font=("bold",20),width=15)
                computer_marks_9th.place(x=1270,y=440)
                ###marks ends
                export_button_entry_var_9th=StringVar()
                export_button_entry_9th=Entry(student_window_9th,textvariable=export_button_entry_var_9th,font=("bold",20),width=15)
                export_button_entry_9th.place(x=180,y=710)


                
                
                #ok button db connecter starts
                def upload_marks_data_mysql_9th():
                    #get values of details
                    main_data_name_9th=name_label_9th_var.get()
                    main_data_rollno_9th=rollno_label_9th_var.get()
                    main_data_section_9th=section_label_9th_var.get()
                    main_data_exam_9th=exam_label_9th_var.get()
                    local_class_9th='IX'
                    #get values of marks
                    main_data_english_9th=english_marks_var_9th.get()
                    main_data_t_h_9th=t_h_marks_var_9th.get()
                    main_data_maths_9th=maths_marks_var_9th.get()
                    main_data_science_9th=science_marks_var_9th.get()
                    main_data_social_9th=social_marks_var_9th.get()
                    main_data_computer_9th=computer_marks_var_9th.get()
                    #hindi/telugu check var
                    check_t_h_9th=t_h_var_9th.get()
                    
                    table_name_t_9th=log_username_get + '_9th_telugu'
                    table_name_h_9th=log_username_get + '_9th_hindi'
                    
                    if main_data_name_9th and main_data_rollno_9th and main_data_section_9th and main_data_exam_9th and main_data_english_9th and main_data_t_h_9th and main_data_maths_9th and main_data_science_9th and main_data_social_9th and main_data_computer_9th != None:
                        if check_t_h_9th == 'Telugu':
                            query_t_9th=f'create table if not exists {table_name_t_9th}(rollno int(3) primary key,name varchar(50) not null,class varchar(4) not null,section char(1) not null,examination varchar(30) not null,english varchar(3) not null,telugu varchar(3) not null,mathematics varchar(3) not null,general_science varchar(3) not null,social_science varchar(3) not null,computer_science varchar(3) not null)'
                            cursor.execute(query_t_9th)
                            update_t_9th=f'insert into {table_name_t_9th} values({main_data_rollno_9th},"{main_data_name_9th}","{local_class_9th}","{main_data_section_9th}","{main_data_exam_9th}",{main_data_english_9th},{main_data_t_h_9th},{main_data_maths_9th},{main_data_science_9th},{main_data_social_9th},{main_data_computer_9th})'
                            cursor.execute(update_t_9th)
                            mydb.commit()
                        elif check_t_h_9th == 'Hindi':
                            query_h_9th=f'create table if not exists {table_name_h_9th}(rollno int(3) primary key,name varchar(30) not null,class varchar(4) not null,section char(1) not null,examination varchar(30) not null,english varchar(3) not null,hindi varchar(3) not null,mathematics varchar(3) not null,general_science varchar(3) not null,social_science varchar(3) not null,computer_science varchar(3) not null)'
                            cursor.execute(query_h_9th)
                            update_h_9th=f'insert into {table_name_h_9th} values({main_data_rollno_9th},"{main_data_name_9th}","{local_class_9th}","{main_data_section_9th}","{main_data_exam_9th}",{main_data_english_9th},{main_data_t_h_9th},{main_data_maths_9th},{main_data_science_9th},{main_data_social_9th},{main_data_computer_9th})'
                            cursor.execute(update_h_9th)
                            mydb.commit()
                    else:
                            msg.showwarning("Warning","please fill all details")
                            
                    pass
                upload_marks_data_9th=Button(student_window_9th,text='OK',font=("bold",20),command=upload_marks_data_mysql_9th)
                upload_marks_data_9th.place(x=1400,y=510)
                #ok button db connecter ends
                #####database entry ends

                #####export data starts 9th
                export_label_9th=Label(student_window_9th,text='EXPORT',foreground='red',font=("bold",20))
                export_label_9th.place(x=20,y=610)

                #class selection starts
                export_class_label_9th=Label(student_window_9th,text='Class',font=("bold",20))
                export_class_label_9th.place(x=20,y=650)

                export_class_var_9th=StringVar()
                export_class_entry_9th=ttk.Combobox(student_window_9th,textvariable=export_class_var_9th,font=("bold",20),width=6)
                export_class_entry_9th.place(x=160,y=650)
                export_class_entry_9th['values']=('6th','7th','8th','9th','10th','11th','12th')
                #class selection ends


                #section selection starts
                export_section_label_9th=Label(student_window_9th,text='Section',font=("bold",20))
                export_section_label_9th.place(x=300,y=650)

                export_section_var_9th=StringVar()
                export_section_entry_9th=ttk.Combobox(student_window_9th,textvariable=export_section_var_9th,font=("bold",20),width=6)
                export_section_entry_9th.place(x=440,y=650)
                export_section_entry_9th['values']=('A','B','C')
                #section selection ends

                #optional_sub selection starts
                export_optional_sub_label_9th=Label(student_window_9th,text='Optional Subject',font=("bold",20))
                export_optional_sub_label_9th.place(x=580,y=650)

                export_optional_sub_var_9th=StringVar()
                export_optional_sub_entry_9th=ttk.Combobox(student_window_9th,textvariable=export_optional_sub_var_9th,font=("bold",20),width=6)
                export_optional_sub_entry_9th.place(x=820,y=650)
                export_optional_sub_entry_9th['values']=('telugu','hindi')
                #optional_sub selection ends

                #exam selection starts
                export_exam_label_9th=Label(student_window_9th,text='Examination',font=("bold",20))
                export_exam_label_9th.place(x=1000,y=650)
                
                export_exam_var_9th=StringVar()
                export_exam_entry_9th=ttk.Combobox(student_window_9th,textvariable=export_exam_var_9th,font=("bold",20),width=15)
                export_exam_entry_9th.place(x=1190,y=650)
                export_exam_entry_9th['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                #exam selection ends


                #export button starts
                def export_9th():
                    exp_cls_9th=export_class_var_9th.get()
                    exp_sec_9th=export_section_var_9th.get()
                    exp_opt_sub_9th=export_optional_sub_var_9th.get()
                    exp_exami_9th=export_exam_var_9th.get()
                    export_name_9th_get=export_button_entry_var_9th.get()
                    print(exp_exami_9th)
                    error_fix_9th='error'
                    if exp_cls_9th and exp_sec_9th and exp_opt_sub_9th  and exp_exami_9th and export_name_9th_get and error_fix_9th != None:
                        
                        class_list_9th=['6th','7th','8th','9th','10th','11th','12th']
                        for i_9th in class_list_9th:
                            if exp_cls_9th == i_9th:
                                opt_sub_tem_9th= f"{exp_opt_sub_9th}"
                                print(opt_sub_tem_9th)
                                temp_cls_tab_9th=log_username_get + '_' + i_9th + '_' + str(opt_sub_tem_9th)
                                cursor.execute(f'select * from {temp_cls_tab_9th} where section = "{exp_sec_9th}" and examination = "{exp_exami_9th}"')
                                get_exp_9th=cursor.fetchall()
                                
                                print(get_exp_9th)
                                wb_9th=Workbook()
                                sheet_9th=wb_9th.active
                                half_9th=["rollno","name","class","section","examination","english"]
                                half_9th.append(opt_sub_tem_9th)
                                full_9th=["mathematics","general_science","social_science","computer_science"]
                                complete_list_9th=half_9th + full_9th
                                print(type(complete_list_9th))
                                sheet_9th.append(complete_list_9th)
        
                                
                                for ex_i_9th in get_exp_9th:
                                    sheet_9th.append(ex_i_9th)
                                export_name_9th_get=export_button_entry_var_9th.get()
                                export_filename_get_9th=filedialog.askdirectory()
                                savespot_9th=str(export_filename_get_9th)

                                wb_9th.save(savespot_9th + '/' + str(export_name_9th_get) + '.xlsx')
                                print(savespot_9th + str(export_name_9th_get) + '.xlsx')


                    else:
                        msg.showwarning('Warning','Please fill all details')
                    pass
                export_button_9th=Button(student_window_9th,text='Save AS',command=export_9th,font=('bold',20))
                export_button_9th.place(x=20,y=700)
                #export button ends


                
                #####export data ends 9th
                student_window_9th.geometry('1920x1080')
                student_window_9th.mainloop()
            #student data ENDS

            dash_9th=Button(dash,text='IX',font=('bold',30),width=20,command = _9th_student)
            dash_9th.place(x=20,y=450)


            def _10th_student():            
                student_window_10th = Toplevel()
                student_window_10th.title('STUDENT DATA ENTRY')
                #frame label
                frame_label_10th=Label(student_window_10th,text='',width=212,relief='groove',height=11,bd=20)
                frame_label_10th.place(x=0,y=0)

                frame_label_data_10th=Label(student_window_10th,text='',width=212,relief='groove',height=24,bd=20)
                frame_label_data_10th.place(x=0,y=200)

                frame_label_export_10th=Label(student_window_10th,text='',width=212,relief='groove',height=11,bd=20)
                frame_label_export_10th.place(x=0,y=600)
                #
                
                open_image_10th=Image.open(photo_fetch_address)
                dash_photo_10th=ImageTk.PhotoImage(open_image)
                photo_label_10th=Label(student_window_10th,image=dash_photo_10th)
                photo_label_10th.place(x=20,y=20)


                dash_school_label_10th=Label(student_window_10th,text=school_fetch_name,font=("Algerian",40))
                dash_school_label_10th.place(x=150,y=60)
                #####database entry starts


                #roll no of student starts
                rollno_label_10th=Label(student_window_10th,text='Roll No',font=("bold",20))
                rollno_label_10th.place(x=20,y=230)
                
                rollno_label_10th_var=IntVar()
                rollno_label_10th_entry=Entry(student_window_10th,textvariable=rollno_label_10th_var,width=20,font=("bold",20))
                rollno_label_10th_entry.place(x=180,y=230)
                #roll no of student ends


                #name of student starts
                name_label_10th=Label(student_window_10th,text='Name',font=("bold",20))
                name_label_10th.place(x=20,y=300)#+70
                
                name_label_10th_var=StringVar()
                name_label_10th_entry=Entry(student_window_10th,textvariable=name_label_10th_var,width=20,font=("bold",20))
                name_label_10th_entry.place(x=180,y=300)
                #name of student ends

                #section of student starts
                section_label_10th=Label(student_window_10th,text='Section',font=("bold",20))
                section_label_10th.place(x=600,y=230)#+70
                
                section_label_10th_var=StringVar()
                section_label_10th_combo=ttk.Combobox(student_window_10th,textvariable=section_label_10th_var,width=20,font=("bold",20))
                section_label_10th_combo['values']=('A','B','C')
                section_label_10th_combo.current(0)
                section_label_10th_combo.place(x=760,y=230)
                #section of student ends


                #exam of student starts
                exam_label_10th=Label(student_window_10th,text='Examination',font=("bold",20))
                exam_label_10th.place(x=600,y=300)#+70
                
                exam_label_10th_var=StringVar()
                exam_label_10th_combo=ttk.Combobox(student_window_10th,textvariable=exam_label_10th_var,width=20,font=("bold",20))
                exam_label_10th_combo['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                exam_label_10th_combo.current(0)
                exam_label_10th_combo.place(x=760,y=300)
                #exam of student ends
                

                ###subjects start
                subject_10th=Label(student_window_10th,text='Subject',font=("bold",20),foreground='#008f00')
                subject_10th.place(x=20,y=370)

                english_10th=Label(student_window_10th,text='English',font=("bold",20))
                english_10th.place(x=180,y=370)

                t_h_var_10th=StringVar()
                t_h_10th=ttk.Combobox(student_window_10th,textvariable=t_h_var_10th,font=("bold",20),width=6)
                t_h_10th['values']=('Telugu','Hindi')
                t_h_10th.current(0)
                t_h_10th.place(x=340,y=370)

                maths_10th=Label(student_window_10th,text='Mathematics',font=("bold",20))
                maths_10th.place(x=500,y=370)

                science_10th=Label(student_window_10th,text='General Science',font=("bold",20))
                science_10th.place(x=730,y=370)

                social_10th=Label(student_window_10th,text='Social Science',font=("bold",20))
                social_10th.place(x=1020,y=370)

                computer_10th=Label(student_window_10th,text='Computer Science',font=("bold",20))
                computer_10th.place(x=1270,y=370)

                ###subjects ends


                ###marks start
                marks_10th=Label(student_window_10th,text='Marks',font=("bold",20),foreground='#008f00')
                marks_10th.place(x=20,y=440)

                english_marks_var_10th=IntVar()
                english_marks_10th=Entry(student_window_10th,textvariable=english_marks_var_10th,font=("bold",20),width=6)
                english_marks_10th.place(x=180,y=440)

                t_h_marks_var_10th=IntVar()
                t_h_marks_10th=Entry(student_window_10th,textvariable=t_h_marks_var_10th,font=("bold",20),width=7)
                t_h_marks_10th.place(x=340,y=440)

                maths_marks_var_10th=IntVar()
                maths_marks_10th=Entry(student_window_10th,textvariable=maths_marks_var_10th,font=("bold",20),width=10)
                maths_marks_10th.place(x=500,y=440)

                science_marks_var_10th=IntVar()
                science_marks_10th=Entry(student_window_10th,textvariable=science_marks_var_10th,font=("bold",20),width=13)
                science_marks_10th.place(x=730,y=440)

                social_marks_var_10th=IntVar()
                social_marks_10th=Entry(student_window_10th,textvariable=social_marks_var_10th,font=("bold",20),width=12)
                social_marks_10th.place(x=1020,y=440)

                computer_marks_var_10th=IntVar()
                computer_marks_10th=Entry(student_window_10th,textvariable=computer_marks_var_10th,font=("bold",20),width=15)
                computer_marks_10th.place(x=1270,y=440)
                ###marks ends
                export_button_entry_var_10th=StringVar()
                export_button_entry_10th=Entry(student_window_10th,textvariable=export_button_entry_var_10th,font=("bold",20),width=15)
                export_button_entry_10th.place(x=180,y=710)


                
                
                #ok button db connecter starts
                def upload_marks_data_mysql_10th():
                    #get values of details
                    main_data_name_10th=name_label_10th_var.get()
                    main_data_rollno_10th=rollno_label_10th_var.get()
                    main_data_section_10th=section_label_10th_var.get()
                    main_data_exam_10th=exam_label_10th_var.get()
                    local_class_10th='X'
                    #get values of marks
                    main_data_english_10th=english_marks_var_10th.get()
                    main_data_t_h_10th=t_h_marks_var_10th.get()
                    main_data_maths_10th=maths_marks_var_10th.get()
                    main_data_science_10th=science_marks_var_10th.get()
                    main_data_social_10th=social_marks_var_10th.get()
                    main_data_computer_10th=computer_marks_var_10th.get()
                    #hindi/telugu check var
                    check_t_h_10th=t_h_var_10th.get()
                    
                    table_name_t_10th=log_username_get + '_10th_telugu'
                    table_name_h_10th=log_username_get + '_10th_hindi'
                    
                    if main_data_name_10th and main_data_rollno_10th and main_data_section_10th and main_data_exam_10th and main_data_english_10th and main_data_t_h_10th and main_data_maths_10th and main_data_science_10th and main_data_social_10th and main_data_computer_10th != None:
                        if check_t_h_10th == 'Telugu':
                            query_t_10th=f'create table if not exists {table_name_t_10th}(rollno int(3) primary key,name varchar(50) not null,class varchar(4) not null,section char(1) not null,examination varchar(30) not null,english varchar(3) not null,telugu varchar(3) not null,mathematics varchar(3) not null,general_science varchar(3) not null,social_science varchar(3) not null,computer_science varchar(3) not null)'
                            cursor.execute(query_t_10th)
                            update_t_10th=f'insert into {table_name_t_10th} values({main_data_rollno_10th},"{main_data_name_10th}","{local_class_10th}","{main_data_section_10th}","{main_data_exam_10th}",{main_data_english_10th},{main_data_t_h_10th},{main_data_maths_10th},{main_data_science_10th},{main_data_social_10th},{main_data_computer_10th})'
                            cursor.execute(update_t_10th)
                            mydb.commit()
                        elif check_t_h_10th == 'Hindi':
                            query_h_10th=f'create table if not exists {table_name_h_10th}(rollno int(3) primary key,name varchar(30) not null,class varchar(4) not null,section char(1) not null,examination varchar(30) not null,english varchar(3) not null,hindi varchar(3) not null,mathematics varchar(3) not null,general_science varchar(3) not null,social_science varchar(3) not null,computer_science varchar(3) not null)'
                            cursor.execute(query_h_10th)
                            update_h_10th=f'insert into {table_name_h_10th} values({main_data_rollno_10th},"{main_data_name_10th}","{local_class_10th}","{main_data_section_10th}","{main_data_exam_10th}",{main_data_english_10th},{main_data_t_h_10th},{main_data_maths_10th},{main_data_science_10th},{main_data_social_10th},{main_data_computer_10th})'
                            cursor.execute(update_h_10th)
                            mydb.commit()
                    else:
                            msg.showwarning("Warning","please fill all details")
                            
                    pass
                upload_marks_data_10th=Button(student_window_10th,text='OK',font=("bold",20),command=upload_marks_data_mysql_10th)
                upload_marks_data_10th.place(x=1400,y=510)
                #ok button db connecter ends
                #####database entry ends

                #####export data starts 10th
                export_label_10th=Label(student_window_10th,text='EXPORT',foreground='red',font=("bold",20))
                export_label_10th.place(x=20,y=610)

                #class selection starts
                export_class_label_10th=Label(student_window_10th,text='Class',font=("bold",20))
                export_class_label_10th.place(x=20,y=650)

                export_class_var_10th=StringVar()
                export_class_entry_10th=ttk.Combobox(student_window_10th,textvariable=export_class_var_10th,font=("bold",20),width=6)
                export_class_entry_10th.place(x=160,y=650)
                export_class_entry_10th['values']=('6th','7th','8th','9th','10th','11th','12th')
                #class selection ends


                #section selection starts
                export_section_label_10th=Label(student_window_10th,text='Section',font=("bold",20))
                export_section_label_10th.place(x=300,y=650)

                export_section_var_10th=StringVar()
                export_section_entry_10th=ttk.Combobox(student_window_10th,textvariable=export_section_var_10th,font=("bold",20),width=6)
                export_section_entry_10th.place(x=440,y=650)
                export_section_entry_10th['values']=('A','B','C')
                #section selection ends

                #optional_sub selection starts
                export_optional_sub_label_10th=Label(student_window_10th,text='Optional Subject',font=("bold",20))
                export_optional_sub_label_10th.place(x=580,y=650)

                export_optional_sub_var_10th=StringVar()
                export_optional_sub_entry_10th=ttk.Combobox(student_window_10th,textvariable=export_optional_sub_var_10th,font=("bold",20),width=6)
                export_optional_sub_entry_10th.place(x=820,y=650)
                export_optional_sub_entry_10th['values']=('telugu','hindi')
                #optional_sub selection ends

                #exam selection starts
                export_exam_label_10th=Label(student_window_10th,text='Examination',font=("bold",20))
                export_exam_label_10th.place(x=1000,y=650)
                
                export_exam_var_10th=StringVar()
                export_exam_entry_10th=ttk.Combobox(student_window_10th,textvariable=export_exam_var_10th,font=("bold",20),width=15)
                export_exam_entry_10th.place(x=1190,y=650)
                export_exam_entry_10th['values']=('PRE-MIDTERM','MIDTERM','POST-MIDTERM','ANNUAL')
                #exam selection ends


                #export button starts
                def export_10th():
                    exp_cls_10th=export_class_var_10th.get()
                    exp_sec_10th=export_section_var_10th.get()
                    exp_opt_sub_10th=export_optional_sub_var_10th.get()
                    exp_exami_10th=export_exam_var_10th.get()
                    export_name_10th_get=export_button_entry_var_10th.get()
                    print(exp_exami_10th)
                    error_fix_10th='error'
                    if exp_cls_10th and exp_sec_10th and exp_opt_sub_10th  and exp_exami_10th and export_name_10th_get and error_fix_10th != None:
                        
                        class_list_10th=['6th','7th','8th','9th','10th','11th','12th']
                        for i_10th in class_list_10th:
                            if exp_cls_10th == i_10th:
                                opt_sub_tem_10th= f"{exp_opt_sub_10th}"
                                print(opt_sub_tem_10th)
                                temp_cls_tab_10th=log_username_get + '_' + i_10th + '_' + str(opt_sub_tem_10th)
                                cursor.execute(f'select * from {temp_cls_tab_10th} where section = "{exp_sec_10th}" and examination = "{exp_exami_10th}"')
                                get_exp_10th=cursor.fetchall()
                                
                                print(get_exp_10th)
                                wb_10th=Workbook()
                                sheet_10th=wb_10th.active
                                half_10th=["rollno","name","class","section","examination","english"]
                                half_10th.append(opt_sub_tem_10th)
                                full_10th=["mathematics","general_science","social_science","computer_science"]
                                complete_list_10th=half_10th + full_10th
                                print(type(complete_list_10th))
                                sheet_10th.append(complete_list_10th)
        
                                
                                for ex_i_10th in get_exp_10th:
                                    sheet_10th.append(ex_i_10th)
                                export_name_10th_get=export_button_entry_var_10th.get()
                                export_filename_get_10th=filedialog.askdirectory()
                                savespot_10th=str(export_filename_get_10th)

                                wb_10th.save(savespot_10th + '/' + str(export_name_10th_get) + '.xlsx')
                                print(savespot_10th + str(export_name_10th_get) + '.xlsx')


                    else:
                        msg.showwarning('Warning','Please fill all details')
                    pass
                export_button_10th=Button(student_window_10th,text='Save AS',command=export_10th,font=('bold',20))
                export_button_10th.place(x=20,y=700)
                #export button ends


                
                #####export data ends 10th
                student_window_10th.geometry('1920x1080')
                student_window_10th.mainloop()
            #student data ENDS

            dash_10th=Button(dash,text='X',font=('bold',30),width=20,command = _10th_student)
            dash_10th.place(x=20,y=530)

            dash_11th=Button(dash,text='XI',font=('bold',30),width=20)
            dash_11th.place(x=20,y=610)

            dash_12th=Button(dash,text='XII',font=('bold',30),width=20)
            dash_12th.place(x=20,y=690)
            
            dash.geometry("1920x1080")
            dash.mainloop()
       
        
    win2_reg_button=Button(win2,text="LOGIN",command=log2_click,bg="red",width=10,font=("bold",30),relief="raised")
    win2_reg_button.place(x=600,y=520)
    win2.geometry("1920x1080")
    win2.mainloop()

win1_log_button=Button(win1,text="Already Member? Login",bg="#FDFEFE",foreground="blue",command=log1_click,relief="raised")
win1_log_button.place(x=800,y=680)
#register_cloud

def upload_photo():
    global photo_address
    file=filedialog.askopenfilename()
    photo_address=file
    
upload_photo_button=Button(win1,text="UPLOAD",font=("bold",30),command=upload_photo)
upload_photo_button.place(x=750,y=580)

def reg_click():
    reg_username_get=username_win1_entry.get()
    reg_email_get=email_win1_entry.get()
    reg_pass_get=pass_win1_entry.get()
    reg_school_get=school_win1_entry.get()
    sql_1="insert into users(username,email,password,schoolname,profilephoto) values(%s,%s,%s,%s,%s)"
    val_1=(reg_username_get,reg_email_get,reg_pass_get,reg_school_get,photo_address)
    cursor.execute(sql_1,val_1)
    mydb.commit()
    msg.showinfo("GOOD NEWS","Sucessfully Registered")
win1_reg_button=Button(win1,text="REGISTER",command=reg_click,bg="red",width=10,font=("bold",30),relief="raised")
win1_reg_button.place(x=600,y=710)

try:
    mydb=conn.connect(host='bc7xufc1ikcn1wzxvahd-mysql.services.clever-cloud.com',user='u9satynarcpueetk',password='VccKjRLxbo5xUmpblg0R',database='bc7xufc1ikcn1wzxvahd')
    cursor=mydb.cursor()
except conn.Error as err:
    msg.showwarning("Warning","check your Internet Connection")
    win1.destroy()

#Upload photo starts
photo_win1_label=Label(win1,text="YOUR PHOTO",width=12,bg="#3E9DE4",foreground="#FDFEFE",font=("Algerian",40),relief="raised",borderwidth=10)
photo_win1_label.place(x=250,y=580)


#Upload photo ends
win1.geometry("1920x1080")
win1.mainloop()
